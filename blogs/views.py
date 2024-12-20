import json
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views import View
from django.db.models import F
from django.urls import reverse
from django.db.models import Q, Sum
from django.views.generic import TemplateView
from django.conf import settings
from accounts.models import DatosTarjeta
from django.utils.timezone import make_aware,localtime, is_naive
from datetime import datetime, timedelta
from django.utils import timezone
from .models import *
from .models import Notification, Category, FavoriteCategory, Blog, Rating
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from django.views.decorators.csrf import csrf_exempt
from PIL import Image as PILImage
import base64
from io import BytesIO
import stripe
stripe.api_key = settings.STRIPE_SECRET_KEY
import requests
from django.db.models import Count

"""
Este módulo define las vistas para la aplicación de blogs, incluyendo la visualización de blogs, creación de comentarios, respuestas, marcadores y likes.

Funciones:
- get_blog_url(slug): Devuelve la URL de un blog basado en su slug.

Clases:
- BlogView: Vista para mostrar una entrada de blog.
- CreateBookmark: Vista para crear o eliminar un marcador en un blog.
- CreateLike: Vista para crear o eliminar un like en un blog.
- CreateComment: Vista para crear un comentario en un blog.
- CreateReply: Vista para crear una respuesta a un comentario.
"""

def get_blog_url(slug):
    """
    Devuelve la URL de un blog basado en su slug.
    """
    return reverse("blogs:blog", args=[slug])

class BlogView(View):
    """
    Vista para mostrar una entrada de blog.
    """
    def get(self, request, slug):
        """
        Maneja las solicitudes GET para mostrar una entrada de blog específica.
        """
        queryset = Blog.objects.filter(is_active=True, slug=slug)
        blog = get_object_or_404(queryset)
      

        category = blog.category

        if category.subcategory_type == "paga":
            if request.user.is_authenticated:
                has_membership = PaidMembership.objects.filter(user=request.user, category=category).exists()
                if not has_membership:
                    return redirect('get_category', slug=category.slug)
            else:
                return redirect('get_category', slug=category.slug)

        blog.views = F("views") + 1
        blog.save()
        blog.refresh_from_db()

        context = {
            "blog": blog,
            "comments": blog.comments.filter(is_active=True)
        }
        if request.user.is_authenticated:
            context["bookmarked"] = Bookmark.objects.filter(creator=request.user, blog=blog).first()
            context["liked"] = BlogLike.objects.filter(blog=blog, creator=request.user).first()

        return render(request, "blogs/blog.html", context)

class CreateBookmark(View):
    """
    Vista para crear o eliminar un marcador en un blog.
    """
    def post(self, request):
        """
        Maneja las solicitudes POST para crear o eliminar un marcador en un blog.
        """
        blog = get_object_or_404(Blog.objects.filter(is_active=True, id=request.POST.get("id")))
        bookmarked = Bookmark.objects.filter(creator=request.user, blog=blog).first()
        if bookmarked:
            bookmarked.delete()
            messages.info(request, "Marcador eliminado")
        else:
            b = Bookmark(
                blog = blog,
                creator = request.user
            )
            b.save()
            messages.success(request, "Marcador creado")
        return redirect("blogs:blog", slug=blog.slug)

class CreateLike(View):
    """
    Vista para crear o eliminar un like en un blog.
    """
    def post(self, request):
        """
        Maneja las solicitudes POST para crear o eliminar un like en un blog.
        """
        blog = get_object_or_404(Blog.objects.filter(is_active=True, id=request.POST.get("id")))
        liked = BlogLike.objects.filter(blog=blog, creator=request.user).first()
        if liked:
            liked.delete()
            messages.info(request, "Me gusta eliminado")
        else:
            l = BlogLike(creator=request.user, blog=blog)
            l.save()
            messages.success(request, "Me gusta este blog")
        return redirect('blogs:blog', slug=blog.slug)
    
@method_decorator(csrf_exempt, name='dispatch')
class IncrementShareCountView(View):
    def post(self, request, blog_id):
        try:
            print(f"Solicitud recibida para incrementar el contador de compartidos para blog_id: {blog_id}")
            blog = Blog.objects.get(id=blog_id)
            print(f"Blog encontrado: {blog.title} con share_count actual: {blog.share_count}")
            blog.share_count = F('share_count') + 1
            blog.save()
            blog.refresh_from_db()  # Refrescar el objeto para obtener el valor actualizado
            print(f"Nuevo share_count para blog_id {blog_id}: {blog.share_count}")
            return JsonResponse({'success': True})
        except Blog.DoesNotExist:
            print(f"Blog con id {blog_id} no encontrado")
            return JsonResponse({'success': False, 'error': 'Blog no encontrado'})
        except Exception as e:
            print(f"Ocurrió un error: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        

class CreateComment(View):
    """
    Vista para crear un comentario en un blog.
    """
    def post(self, request):
        """
        Maneja las solicitudes POST para crear un comentario en un blog.
        """
        blog = get_object_or_404(Blog.objects.filter(is_active=True, id=request.POST.get("id")))
        comment = request.POST.get("comment")
        if not comment:
            messages.warning(request, "Comentario requerido")
            return redirect("blogs:blog", slug=blog.slug)
        c = Comment(
            comment = comment,
            creator = request.user,
            blog = blog
        )
        c.save()
        messages.success(request, "Comentario creado")
        return redirect(get_blog_url(blog.slug)+"#comments")

class CreateReply(View):
    """
    Vista para crear una respuesta a un comentario.
    """
    def post(self, request):
        """
        Maneja las solicitudes POST para crear una respuesta a un comentario.
        """
        comment = get_object_or_404(Comment.objects.filter(is_active=True, id=request.POST.get("id")))
        reply = request.POST.get("reply")
        if not reply:
            messages.warning(request, "Respuesta requerida")
            return redirect("blogs:blog", slug=comment.blog.slug)

        r = Reply(
            reply = request.POST.get("reply", ""),
            comment = comment,
            creator = request.user
        )
        r.save()

        messages.success(request, "Respuesta creada")
        return redirect(get_blog_url(comment.blog.slug)+"#comments")
    


@login_required
def notifications(request):
    """
    Vista para mostrar las notificaciones no leídas del usuario.

    Parámetros:
    request (HttpRequest): El objeto de solicitud HTTP.

    Retorna:
    HttpResponse: Renderiza la plantilla 'notifications.html' con las notificaciones no leídas
                  y el conteo de las mismas.

    Funcionalidad:
    - Filtra las notificaciones del usuario autenticado que no han sido leídas.
    - Ordena las notificaciones por fecha de creación en orden descendente.
    - Cuenta el número de notificaciones no leídas.
    - Renderiza la plantilla 'notifications.html' con las notificaciones y su conteo.
    """
    notifications = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    notifications_count = notifications.count()
    return render(request, 'notifications.html', {'notifications': notifications, 'notifications_count': notifications_count})

@login_required
def mark_as_read(request, notification_id):
    """
    Vista para marcar una notificación específica como leída.

    Parámetros:
    request (HttpRequest): El objeto de solicitud HTTP.
    notification_id (int): El ID de la notificación a marcar como leída.

    Retorna:
    HttpResponseRedirect: Redirige a la vista 'manage:kanban'.

    Funcionalidad:
    - Obtiene la notificación especificada por ID y perteneciente al usuario autenticado.
    - Marca la notificación como leída.
    - Guarda los cambios en la base de datos.
    - Redirige a la vista 'manage:kanban'.
    """
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    return redirect('manage:kanban')

@login_required
def mark_all_as_read(request):
    """
    Vista para marcar todas las notificaciones no leídas del usuario como leídas.

    Parámetros:
    request (HttpRequest): El objeto de solicitud HTTP.

    Retorna:
    HttpResponseRedirect: Redirige a la página anterior usando HTTP_REFERER.

    Funcionalidad:
    - Filtra las notificaciones no leídas del usuario autenticado.
    - Marca todas las notificaciones filtradas como leídas.
    - Redirige a la página anterior usando HTTP_REFERER.
    """
    notifications = Notification.objects.filter(user=request.user, is_read=False)
    notifications.update(is_read=True)
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def toggle_favorite_category(request, category_id):
    """
    Vista para añadir o eliminar una categoría de las favoritas del usuario.

    Parámetros:
    request (HttpRequest): El objeto de solicitud HTTP.
    category_id (int): El ID de la categoría a añadir o eliminar de favoritas.

    Retorna:
    JsonResponse: Retorna un JSON con el estado de la operación ('added' o 'removed').

    Funcionalidad:
    - Obtiene la categoría especificada por ID.
    - Intenta obtener o crear una relación de categoría favorita para el usuario autenticado.
    - Si la relación ya existe, la elimina y retorna un estado 'removed'.
    - Si la relación no existe, la crea y retorna un estado 'added'.
    """
    category = get_object_or_404(Category, id=category_id)
    favorite, created = FavoriteCategory.objects.get_or_create(user=request.user, category=category)
    if not created:
        favorite.delete()
        return JsonResponse({'status': 'removed'})
    return JsonResponse({'status': 'added'})

@login_required
def favorite_categories(request):
    """
    Vista para mostrar las categorías favoritas del usuario.

    Parámetros:
    request (HttpRequest): El objeto de solicitud HTTP.

    Retorna:
    HttpResponse: Renderiza la plantilla 'blogs/favorite_categories.html' con las categorías favoritas.

    Funcionalidad:
    - Filtra las categorías favoritas del usuario autenticado.
    - Renderiza la plantilla 'blogs/favorite_categories.html' con las categorías favoritas.
    """
    favorites = FavoriteCategory.objects.filter(user=request.user)
    return render(request, 'blogs/favorite_categories.html', {'favorites': favorites})


class PagoView(View):
    """
    Vista para manejar el pago de una categoría específica.

    Método:
        get(request, category_id):
            Procesa la solicitud GET para mostrar la página de pago.
    """

    def get(self, request, category_id):
        """
        Maneja la solicitud GET para mostrar el formulario de pago.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            category_id (int): ID de la categoría para el pago.

        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/pago.html' con los datos necesarios.
        """
        category = get_object_or_404(Category, id=category_id)
        datos_tarjeta = None
        if hasattr(request.user, 'datos_tarjeta'):
            datos_tarjeta = {
                'nombre_tarjeta': request.user.datos_tarjeta.nombre_tarjeta,
                'numero_tarjeta': request.user.datos_tarjeta.numero_tarjeta,
                'fecha_vencimiento': request.user.datos_tarjeta.fecha_vencimiento,
                'codigo_seguridad': request.user.datos_tarjeta.codigo_seguridad,
            }
        return render(request, 'blogs/pago.html', {
            'category': category,
            'datos_tarjeta': datos_tarjeta,
            'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY
        })


class CreateCheckoutSessionView(View):
    """
    Vista para crear una sesión de pago con Stripe.

    Método:
        post(request, category_id, *args, **kwargs):
            Procesa la solicitud POST para crear una sesión de pago.
    """

    def post(self, request, category_id, *args, **kwargs):
        """
        Maneja la solicitud POST para crear una sesión de pago mediante Stripe.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            category_id (int): ID de la categoría para el pago.
            *args: Argumentos adicionales.
            **kwargs: Argumentos de palabra clave adicionales.

        Returns:
            HttpResponseRedirect: Redirige al usuario a la URL de la sesión de pago.
            str: Retorna el error en caso de excepción.
        """
        category = get_object_or_404(Category, id=category_id)
        YOUR_DOMAIN = "http://127.0.0.1:8000/"
        try:
            checkout_session = stripe.checkout.Session.create(
                payment_method_types=['card'],
                line_items=[
                    {
                        'price_data': {
                            'currency': 'pyg',
                            'product_data': {
                                'name': category.category,
                                'description': category.desc,
                            },
                            'unit_amount': category.costo_membresia
                        },
                        'quantity': 1,
                    },
                ],
                mode='payment',
                success_url=YOUR_DOMAIN + reverse('blogs:success', kwargs={'category_id': category_id}),
                cancel_url=YOUR_DOMAIN + reverse('blogs:cancel', kwargs={'category_id': category_id}),
            )
            return redirect(checkout_session.url, code=303)
        except Exception as e:
            return str(e)

class SuccessView(View):
    """
    Vista para manejar la página de éxito después de un pago exitoso.

    Método:
        get(request, category_id):
            Procesa la solicitud GET para mostrar la página de éxito.
    """

    def get(self, request, category_id):
        """
        Maneja la solicitud GET para mostrar la página de éxito de pago.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            category_id (int): ID de la categoría para la cual se realizó el pago.

        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/success.html' con los datos necesarios.
        """
        category = get_object_or_404(Category, id=category_id)
        return render(request, 'blogs/success.html', {'category': category})


class CancelView(View):
    """
    Vista para manejar la cancelación de un pago.

    Método:
        get(request, category_id):
            Procesa la solicitud GET para redirigir al usuario tras la cancelación.
    """

    def get(self, request, category_id):
        """
        Maneja la solicitud GET para redirigir al usuario después de la cancelación del pago.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            category_id (int): ID de la categoría para la cual se intentó el pago.

        Returns:
            HttpResponseRedirect: Redirige al usuario a la página de inicio.
        """
        category = get_object_or_404(Category, id=category_id)
        return redirect('index')


@method_decorator(login_required, name='dispatch')
class IrACategoriaView(View):
    """
    Vista para redirigir al usuario a una categoría específica y gestionar la membresía.
    """

    def get(self, request, category_id):
        """
        Maneja la solicitud GET para redirigir al usuario a la categoría seleccionada.
        Verifica si el usuario ya tiene una membresía para la categoría y actúa en consecuencia.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            category_id (int): ID de la categoría a la que se redirige.

        Returns:
            HttpResponseRedirect: Redirige al usuario a la página de la categoría o muestra un mensaje de éxito.
        """
        category = get_object_or_404(Category, id=category_id)

        # Verificar si el usuario ya tiene una membresía para esta categoría
        if PaidMembership.objects.filter(user=request.user, category=category).exists():
            messages.success(request, "Ya tienes una membresía para esta categoría. Ve y disfruta de los Artículos.")
           
        else:
            # Guardar la membresía pagada
            PaidMembership.objects.create(
                user=request.user,
                category=category,
                category_desc=category.desc,
                category_type=category.category_type,
                subcategory_type=category.subcategory_type,
                membership_cost=category.costo_membresia
            )

            # Guardar el registro de pago de la membresía
            MembershipPayment.objects.create(
                user=request.user,
                category=category,
                category_type=category.category_type,
                membership_cost=category.costo_membresia
            )
            messages.success(request, "Tu membresía ha sido activada. Ahora puedes disfrutar de los Artículos.")

        # Redirigir a la página de la categoría
        return redirect('get_category', slug=category.slug)
    

@method_decorator(login_required, name='dispatch')
class MembershipsView(View):
    """
    Vista para gestionar las membresías pagadas del usuario.
    """

    def get(self, request):
        """
        Maneja la solicitud GET para mostrar las membresías del usuario con opciones de filtrado.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.

        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/memberships.html' con las membresías filtradas y el total pagado.
        """
        query = request.GET.get('q')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        memberships = PaidMembership.objects.filter(user=request.user).order_by('-payment_date')

        if query:
            memberships = memberships.filter(
                Q(category__category__icontains=query) |
                Q(category_desc__icontains=query) |
                Q(category_type__iexact=query) |
                Q(membership_cost__icontains=query)
            )

        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        else:
            start_date = None

        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)) - timedelta(microseconds=1)
            if start_date:
                memberships = memberships.filter(payment_date__range=[start_date, end_date])
            else:
                memberships = memberships.filter(payment_date__lte=end_date)
        elif start_date:
            memberships = memberships.filter(payment_date__gte=start_date)

        # Si no se encuentran pagos en el rango de fechas, ajustar start_date a la fecha del primer pago
        if not memberships.exists() and not start_date:
            first_payment = PaidMembership.objects.filter(user=request.user).order_by('payment_date').first()
            if first_payment:
                start_date = first_payment.payment_date
                memberships = PaidMembership.objects.filter(user=request.user, payment_date__gte=start_date)
 

        total_paid = memberships.aggregate(Sum('membership_cost'))['membership_cost__sum'] or 0

        return render(request, 'blogs/memberships.html', {'memberships': memberships, 'total_paid': total_paid})
    

    def post(self, request):
        """
        Maneja la solicitud POST para eliminar una membresía del usuario.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.

        Returns:
            HttpResponseRedirect: Redirige al usuario a la página de membresías después de la eliminación.
        """
        membership_id = request.POST.get('membership_id')
        membership = get_object_or_404(PaidMembership, id=membership_id, user=request.user)
        
        # Eliminar el registro correspondiente en MembershipPayment
        MembershipPayment.objects.filter(user=request.user, category=membership.category).delete()
        
        # Eliminar la membresía
        membership.delete()

        # Recalcular el total pagado
        memberships = PaidMembership.objects.filter(user=request.user)
        total_paid = memberships.aggregate(Sum('membership_cost'))['membership_cost__sum'] or 0

        return redirect('blogs:memberships')

@method_decorator(login_required, name='dispatch')
class ExportMembershipsView(View):
    """
    Vista para exportar las membresías pagadas del usuario a un archivo Excel.
    """

    def get(self, request):
        """
        Maneja la solicitud GET para generar y descargar un archivo Excel con las membresías pagadas.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.

        Returns:
            HttpResponse: Respuesta HTTP con el archivo Excel adjunto.
        """
        query = request.GET.get('q')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        total_paid = request.GET.get('total_paid')
        memberships = PaidMembership.objects.filter(user=request.user).order_by('-payment_date')

        if query:
            memberships = memberships.filter(
                Q(category__category__icontains=query) |
                Q(category_desc__icontains=query) |
                Q(category_type__iexact=query) |
                Q(subcategory_type__iexact=query) |
                Q(membership_cost__icontains=query)
            )

        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)) - timedelta(microseconds=1)
                memberships = memberships.filter(payment_date__range=[start_date, end_date])
            else:
                memberships = memberships.filter(payment_date__gte=start_date)

        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Membresías"

        # Establecer el título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "CLOUDCMS"
        title_cell.font = Font(size=20, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Información del usuario
        user_info = [
            ("Usuario:", request.user.username),
            ("Nombre:", request.user.get_full_name()),
            ("Correo:", request.user.email),
        ]

        row = 3
        for label, value in user_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)  # Aplicar negrita a las etiquetas
            ws[f'B{row}'] = value
            row += 1

        # Espacio entre la información del usuario y la tabla de membresías
        row += 1

        ws.append([])
        ws.append(["Membresías Pagadas en CloudCMS:"])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        ws.append([])

        # Escribir los encabezados
        headers = ["Nombre de la Categoría", "Descripción", "Tipo", "Subcategoría", "Costo (Gs.)","Tipo de Pago", "Fecha de Pago"]
        ws.append(headers)

        # Aplicar el estilo de negrita a los encabezados
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

       # Ajustar el ancho de las columnas
        column_widths = [25, 30, 15, 20, 10, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
 
        
        # Definir el estilo de borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar el estilo de borde a los encabezados
        for cell in ws[ws.max_row]:
            cell.border = thin_border


        # Escribir los datos
        for membership in memberships:
            row = [
                membership.category.category,
                membership.category_desc,
                membership.category_type,
                membership.subcategory_type,
                membership.membership_cost,
                "TC/TD",
                membership.payment_date.strftime("%d %b, %Y %H:%M:%S")
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border


        # Escribir el total pagado
        ws.append([])
        ws.append(["Total General Pagado:", f"Gs. {total_paid}"])
        for cell in ws[ws.max_row]:
            cell.border = thin_border

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Membresias_{request.user.username}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

def filter_payments(query, start_date, end_date):
    payments = MembershipPayment.objects.all().order_by('-payment_date')
    
    if query:
        payments = payments.filter(
            Q(user__username__icontains=query) |
            Q(category__category__icontains=query) |
            Q(category_type__iexact=query) |
            Q(membership_cost__icontains=query)
        )

    if start_date:
        start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
    else:
        start_date = None

    if end_date:
        end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)) - timedelta(microseconds=1)
        if start_date:
            payments = payments.filter(payment_date__range=[start_date, end_date])
        else:
            payments = payments.filter(payment_date__lte=end_date)
    elif start_date:
        payments = payments.filter(payment_date__gte=start_date)

    # Si no se encuentran pagos en el rango de fechas, ajustar start_date a la fecha del primer pago
    if not payments.exists() and not start_date:
        first_payment = MembershipPayment.objects.order_by('payment_date').first()
        if first_payment:
            start_date = first_payment.payment_date
            payments = MembershipPayment.objects.filter(payment_date__gte=start_date)

    return payments

@method_decorator(permission_required('accounts.can_view_membership_payments', raise_exception=True), name='dispatch')
class AllMembershipPaymentsView(View):
    def get(self, request):
        query = request.GET.get('q')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        payments = filter_payments(query, start_date, end_date)
        
        total_paid = payments.aggregate(Sum('membership_cost'))['membership_cost__sum'] or 0

        return render(request, 'blogs/all_membership_payments.html', {'payments': payments, 'total_paid': total_paid})
    
    def post(self, request):
        """
        Maneja la solicitud POST para eliminar un pago de membresía.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.

        Returns:
            HttpResponseRedirect: Redirige al usuario a la página de todos los pagos de membresías después de la eliminación.
        """
        payment_id = request.POST.get('payment_id')
        payment = get_object_or_404(MembershipPayment, id=payment_id)
        
        # Eliminar el registro correspondiente en PaidMembership
        PaidMembership.objects.filter(user=payment.user, category=payment.category).delete()
        
        # Eliminar el registro en MembershipPayment
        payment.delete()

        # Recalcular el total pagado
        payments = MembershipPayment.objects.all()
        total_paid = payments.aggregate(Sum('membership_cost'))['membership_cost__sum'] or 0
        
        return redirect('blogs:all_membership_payments')

@method_decorator(login_required, name='dispatch')
class StatisticsView(TemplateView):
    template_name = 'blogs/estadisticas.html'
    
    """
    Vista para mostrar estadísticas de los pagos de membresías.
    """

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener los parámetros de filtro desde el request
        query = self.request.GET.get('q')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Aplicar el filtro en los pagos
        payments = filter_payments(query, start_date, end_date)

        """
        Maneja la solicitud GET para generar y mostrar estadísticas de pagos de membresías.

        Obtiene datos agregados de las membresías pagadas por categoría y por fecha,
        prepara los datos para visualización en gráficos y renderiza la plantilla correspondiente.

        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.

        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/estadisticas.html' con los datos de estadísticas.
        """
        # Obtener los datos de las categorías y los totales comprados por cada categoría
        categories = payments.values('category__category').annotate(total=Sum('membership_cost')).order_by('-total')

        # Preparar los datos para el gráfico de pie
        labels = [category['category__category'] for category in categories]
        data = [category['total'] for category in categories]

        # Obtener los pagos agrupados por fecha
        payments_by_date = payments.values('payment_date__date').annotate(total=Sum('membership_cost')).order_by('payment_date__date')

        # Preparar los datos para el gráfico de barra/linea de tiempo
        dates = [payment['payment_date__date'].strftime("%Y-%m-%d") for payment in payments_by_date]
        totals = [payment['total'] for payment in payments_by_date]

        # Obtener los pagos agrupados por fecha y categoría
        payments_by_date_and_category = payments.values('payment_date__date', 'category__category').annotate(total=Sum('membership_cost')).order_by('payment_date__date')

        # Preparar los datos para el gráfico de línea de tiempo por categoría
        category_data = {}
        for payment in payments_by_date_and_category:
            category = payment['category__category']
            date = payment['payment_date__date'].strftime("%Y-%m-%d")
            total = payment['total']
            if category not in category_data:
                category_data[category] = []
            category_data[category].append({'x': date, 'y': total})

        # Pasar los datos al contexto
        context['labels'] = json.dumps(labels)
        context['data'] = json.dumps(data)
        context['dates'] = json.dumps(dates)
        context['totals'] = json.dumps(totals)
        context['category_data'] = json.dumps(category_data)
        
        return context

@method_decorator(login_required, name='dispatch')
class ExportStatisticsView(View):
    """
    Vista para exportar las estadísticas de membresías a un archivo Excel.
    """

    def post(self, request):
        """
        Maneja la solicitud POST para generar y descargar un archivo Excel con las estadísticas de membresías.
        
        Obtiene los gráficos desde el formulario, crea un archivo Excel con los datos de las membresías pagadas,
        añade los gráficos al archivo y prepara la respuesta HTTP para la descarga.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
        
        Returns:
            HttpResponse: Respuesta HTTP con el archivo Excel adjunto.
        """
        # Obtener los datos de los gráficos desde el formulario
        pie_chart = request.POST.get('pie_chart')
        bar_chart = request.POST.get('bar_chart')
        line_chart = request.POST.get('line_chart')


        # Obtener los filtros de `query`, `start_date`, y `end_date` desde el request
        query = request.POST.get('q')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        # Filtrar los pagos utilizando la función `filter_payments`
        payments = filter_payments(query, start_date, end_date)

        # Crear un nuevo libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Análisis General de Membresías"

        # Establecer el título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "CLOUDCMS"
        title_cell.font = Font(size=20, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Espacio entre el título y la tabla de membresías
        ws.append([])
        ws.append(["Todas las Membresías Pagadas en CloudCMS:"])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        ws.append([])

        # Escribir los encabezados
        headers = ["Usuario", "Nombre de la Categoría", "Tipo", "Costo (Gs.)", "Tipo de Pago", "Fecha de Pago"]
        ws.append(headers)

        # Aplicar el estilo de negrita a los encabezados
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Ajustar el ancho de las columnas
        column_widths = [25, 30, 15, 20, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Definir el estilo de borde
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar el estilo de borde a los encabezados
        for cell in ws[ws.max_row]:
            cell.border = thin_border

        # Escribir los datos de todas las membresías pagadas
        for payment in payments:
            payment_date = payment.payment_date
            if is_naive(payment_date):
                payment_date = make_aware(payment_date)
            row = [
                payment.user.username,
                payment.category.category,
                payment.category_type,
                f"Gs. {payment.membership_cost}",
                "TC/TD",  # Asumiendo que el tipo de pago es siempre TC/TD
                localtime(payment.payment_date).strftime("%d %b, %Y %H:%M:%S")
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border

        # Calcular el total pagado
        total_paid = payments.aggregate(Sum('membership_cost'))['membership_cost__sum'] or 0

        # Escribir el total pagado
        ws.append([])
        ws.append(["Total General Pagado:", f"Gs. {total_paid}"])
        for cell in ws[ws.max_row]:
            cell.border = thin_border


        # Añadir los gráficos al archivo Excel
        if pie_chart:
            pie_image = self.base64_to_image(pie_chart)
            ws.add_image(pie_image, 'I3')

        if bar_chart:
            bar_image = self.base64_to_image(bar_chart)
            ws.add_image(bar_image, 'I40')

        if line_chart:
            line_image = self.base64_to_image(line_chart)
            ws.add_image(line_image, 'I70')

        # Preparar la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=estadisticas_membresias.xlsx'
        wb.save(response)
        return response

    def base64_to_image(self, base64_string):
        """
        Convierte una cadena en base64 a un objeto de imagen compatible con openpyxl.
        
        Args:
            base64_string (str): Cadena de imagen codificada en base64.
        
        Returns:
            openpyxl.drawing.image.Image: Objeto de imagen para insertar en el archivo Excel.
        """
        image_data = base64.b64decode(base64_string.split(',')[1])
        image = PILImage.open(BytesIO(image_data))
        image_io = BytesIO()
        image.save(image_io, format='PNG')
        image_io.seek(0)
        return Image(image_io)
    

class RateBlogView(View):
    """
    Vista para gestionar la calificación de un blog por parte de un usuario.
    
    Permite a un usuario calificar un blog y actualiza las estadísticas de calificaciones correspondientes.
    """

    def post(self, request, blog_id):
        """
        Maneja la solicitud POST para calificar un blog.
        
        Obtiene la calificación enviada por el usuario, actualiza la calificación del blog y redirige al detalle del blog.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            blog_id (int): ID del blog que se va a calificar.
        
        Returns:
            HttpResponseRedirect: Redirige a la página de detalle del blog.
        """
        blog = get_object_or_404(Blog, id=blog_id)
        rating_value = int(request.POST.get('rating'))
        user = request.user

        # Obtener la calificación existente del usuario si existe
        rating, created = Rating.objects.get_or_create(user=user, blog=blog, defaults={'rating': rating_value})

        # Descontar la calificación anterior
        if not created:
            if rating.rating == 1:
                blog.one_star_ratings -= 1
            elif rating.rating == 2:
                blog.two_star_ratings -= 1
            elif rating.rating == 3:
                blog.three_star_ratings -= 1

        # Actualizar la calificación con la nueva
        rating.rating = rating_value
        rating.save()

        # Incrementar la nueva calificación
        if rating_value == 1:
            blog.one_star_ratings += 1
        elif rating_value == 2:
            blog.two_star_ratings += 1
        elif rating_value == 3:
            blog.three_star_ratings += 1

        blog.save()
        return redirect('manage:blog_detail', id=blog.id)
    


class ReportBlogView(View):
    """
    Vista para gestionar el reporte de un blog por parte de un usuario.
    
    Permite a un usuario reportar un blog proporcionando una razón para el reporte.
    """

    def post(self, request, id):
        """
        Maneja la solicitud POST para reportar un blog.
        
        Obtiene la razón del reporte enviada por el usuario, crea un nuevo registro de reporte
        y redirige al detalle del blog reportado.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            id (int): ID del blog que se va a reportar.
        
        Returns:
            HttpResponseRedirect: Redirige a la página de detalle del blog.
        """
        blog = get_object_or_404(Blog, id=id)
        report_reason = request.POST.get('report_reason')
        Report.objects.create(blog=blog, user=request.user, reason=report_reason)
        return redirect('blogs:blog', slug=blog.slug)


class ReportedBlogsView(View):
    """
    Vista para visualizar todos los reportes de blogs.
    
    Permite a los administradores ver una lista de todos los reportes realizados sobre los blogs.
    """

    def get(self, request):
        """
        Maneja la solicitud GET para mostrar todos los reportes de blogs.
        
        Recupera todos los reportes ordenados por fecha de creación descendente y renderiza
        la plantilla correspondiente.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
        
        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/blogs_reportados.html' con los reportes.
        """
        reports = Report.objects.all().order_by('-created_at')
        return render(request, 'blogs/blogs_reportados.html', {'reports': reports})


class VerificarReporteView(View):
    """
    Vista para verificar un reporte específico de un blog.
    
    Permite a los administradores revisar los detalles de un reporte y el blog asociado.
    """

    def get(self, request, id):
        """
        Maneja la solicitud GET para verificar un reporte específico.
        
        Recupera el reporte por su ID y el blog asociado, luego renderiza la plantilla correspondiente.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            id (int): ID del reporte que se va a verificar.
        
        Returns:
            HttpResponse: Renderiza la plantilla 'blogs/verificar_blog_reportado.html' con el blog y el reporte.
        """
        report = get_object_or_404(Report, id=id)
        blog = report.blog
        return render(request, 'blogs/verificar_blog_reportado.html', {'blog': blog, 'report': report})

class ChangeBlogStatusView(View):
    """
    Vista para cambiar el estado de un blog.
    
    Permite actualizar el estado de un blog específico y gestionar las acciones asociadas al cambio de estado.
    """
    
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        """
        Sobrescribe el método dispatch para eximir de la verificación CSRF.
        
        Args:
            *args: Argumentos posicionales.
            **kwargs: Argumentos de palabra clave.
        
        Returns:
            HttpResponse: Respuesta de la superclase dispatch.
        """
        return super().dispatch(*args, **kwargs)

    def post(self, request, blog_id):
        """
        Maneja la solicitud POST para cambiar el estado de un blog.
        
        Obtiene el nuevo estado desde la solicitud, actualiza el estado del blog, elimina reportes si es necesario
        y retorna una respuesta JSON indicando el resultado de la operación.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            blog_id (int): ID del blog que se va a actualizar.
        
        Returns:
            JsonResponse: Respuesta JSON indicando éxito o error de la operación.
        """
        try:
            print(f"Request received for blog_id: {blog_id}")
            blog = Blog.objects.get(id=blog_id)
            data = json.loads(request.body)  # Cargar el cuerpo de la solicitud como JSON
            new_status = data.get('new_status')  # Obtener new_status del JSON
            previous_status = data.get('previous_status')  # Obtener estado previo del JSON
            Comment = None 

            print(f"Data received: new_status={new_status}, previous_status={previous_status}")

            if new_status is None:
                print("new_status es None")
                return JsonResponse({'success': False, 'error': 'new_status es None'})

            new_status = int(new_status)

            if new_status == 2:  # Estado "En edición"
                blog.previous_status = blog.status  # Guardar el estado anterior
                blog.status = new_status
                blog.is_published = False  # Establecer is_published a False
                
                blog.save()

                # Eliminar todos los reportes del blog
                Report.objects.filter(blog=blog).delete()
                print("Blog status updated and reports deleted successfully")
                
                return JsonResponse({'success': True})
            else:
                print(f"Estado no válido: {new_status}")
                return JsonResponse({'success': False, 'error': 'Estado no válido.'})
        except Blog.DoesNotExist:
            print(f"Blog no encontrado: {blog_id}")
            return JsonResponse({'success': False, 'error': 'Blog no encontrado.'})
        except Exception as e:
            print(f"Error al cambiar el estado del blog: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})

class IgnorarReporteView(View):
    """
    Vista para ignorar y eliminar un reporte de un blog.
    
    Permite a los administradores eliminar un reporte específico hecho por un usuario sobre un blog.
    """
    
    def get(self, request, id):
        """
        Maneja la solicitud GET para ignorar y eliminar un reporte de un blog.
        
        Recupera el reporte por su ID, lo elimina y redirige a la lista de reportes.
        
        Args:
            request (HttpRequest): Objeto de la solicitud HTTP.
            id (int): ID del reporte que se va a ignorar y eliminar.
        
        Returns:
            HttpResponseRedirect: Redirige a la página de todos los reportes de blogs.
        """
        report = get_object_or_404(Report, id=id)
        report.delete()
        return redirect('blogs:blogs_reportados')
    


@method_decorator(login_required, name='dispatch')
class BlogStatisticsView(TemplateView):
    """
    Vista para mostrar estadísticas detalladas de los blogs.

    Esta vista permite a los usuarios autenticados ver estadísticas como la cantidad de compartidos,
    calificaciones, visualizaciones y comentarios de los blogs. Además, integra datos de la API de Disqus
    para obtener información sobre los comentarios y presenta los blogs más comentados.

    Atributos:
        template_name (str): Nombre de la plantilla utilizada para renderizar la vista.

    Métodos:
        get_context_data(**kwargs):
            Obtiene y procesa los datos necesarios para mostrar las estadísticas en la plantilla.
    """
    template_name = 'blogs/blog_statistics.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener los parámetros de filtro desde el request
        author_id = self.request.GET.get('author')
        category_id = self.request.GET.get('category')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Filtrar los blogs
        blogs = Blog.objects.filter(is_active=True, is_published=True).order_by('-published_on')
        if author_id:
            blogs = blogs.filter(creator_id=author_id)
        if category_id:
            blogs = blogs.filter(category_id=category_id)
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            blogs = blogs.filter(published_on__gte=start_date)
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)) - timedelta(microseconds=1)
            blogs = blogs.filter(published_on__lte=end_date)
        
        # Obtener los datos de los blogs y los totales de compartidos, calificaciones y visualizaciones
        blog_data = blogs.values('title', 'share_count', 'one_star_ratings', 'two_star_ratings', 'three_star_ratings', 'views')
        
        # Filtrar los datos para cada gráfico
        one_star_data = [blog for blog in blog_data if blog['one_star_ratings'] > 0]
        two_star_data = [blog for blog in blog_data if blog['two_star_ratings'] > 0]
        three_star_data = [blog for blog in blog_data if blog['three_star_ratings'] > 0]

        # Preparar los datos para el gráfico de barras
        labels = [blog['title'] for blog in blog_data]
        share_counts = [blog['share_count'] for blog in blog_data]

        one_star_labels = [blog['title'] for blog in one_star_data]
        one_star_ratings = [blog['one_star_ratings'] for blog in blog_data]

        two_star_labels = [blog['title'] for blog in two_star_data]
        two_star_ratings = [blog['two_star_ratings'] for blog in blog_data]

        three_star_labels = [blog['title'] for blog in three_star_data]
        three_star_ratings = [blog['three_star_ratings'] for blog in blog_data]

        views = [blog['views'] for blog in blog_data]

        # Pasar los datos al contexto
        context['labels'] = json.dumps(labels)
        context['share_counts'] = json.dumps(share_counts)
        context['one_star_labels'] = json.dumps(one_star_labels)
        context['one_star_ratings'] = json.dumps(one_star_ratings)
        context['two_star_labels'] = json.dumps(two_star_labels)
        context['two_star_ratings'] = json.dumps(two_star_ratings)
        context['three_star_labels'] = json.dumps(three_star_labels)
        context['three_star_ratings'] = json.dumps(three_star_ratings)
        context['views'] = json.dumps(views)
        
        # Pasar los autores y categorías al contexto para los filtros
        context['authors'] = User.objects.filter(role='author')
        context['categories'] = Category.objects.all()

        # Configuración de Disqus
        DISQUS_SHORTNAME = 'cloudcms'
        DISQUS_API_KEY = 'CspBCzjGe9AExwI2rz5CsJu43fsw1vExTAfLI09s1W1F0ll5O6Td4G8BAd97TJ7B'
        DISQUS_API_URL = 'https://disqus.com/api/3.0/threads/list.json'

        total_comments = 0
        top_5_blogs = []

        try:
            # Parámetros para la API de Disqus
            params = {
                'forum': DISQUS_SHORTNAME,
                'api_key': DISQUS_API_KEY,
                'limit': 100,
                'include': 'open'
            }

            # Obtener datos de la API
            response = requests.get(DISQUS_API_URL, params=params)
            response.raise_for_status()
            data = response.json()
            threads = data.get('response', [])

            # Procesar datos de Disqus
            blogs_data = []
            for thread in threads:
                if not thread.get('isDeleted', False):  # Excluir hilos eliminados
                    title = thread.get('title', 'Título no disponible')
                    posts = thread.get('posts', 0)
                    link = thread.get('link', 'Enlace no disponible')
                    total_comments += posts  # Acumular comentarios
                    blogs_data.append({'title': title, 'posts': posts, 'link': link})

            # Ordenar por número de comentarios y obtener los 5 más comentados
            top_5_blogs = sorted(blogs_data, key=lambda x: x['posts'], reverse=True)[:5]

        except Exception as e:
            total_comments = 'Error al obtener los comentarios'
            top_5_blogs = []

        # Pasar los datos al contexto
        context['total_comments'] = total_comments
        context['top_5_blogs'] = top_5_blogs

        return context